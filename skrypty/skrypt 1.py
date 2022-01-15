from urllib.request import urlopen

from bs4 import BeautifulSoup

from openpyxl import Workbook

wb = Workbook()

sheet = wb.active

sheet["A1"] = 'kod obi'
 

sheet["B1"]= 'nazwa Obi'

sheet["C1"]='Ceny OBI (stara nowa)'

sheet["D1"] = 'Numer strony'

adresy = ['https://www.obi.pl/campaign/100-zobacz-wiecej']
 

a=[]
b=[]
c=[]

ii=0
i=0
ad=0

for adres in adresy:

    ad=+1
    html= urlopen(adres)
    bs = BeautifulSoup(html.read(),'html.parser')
    nameList1 = bs.find_all('input',attrs={'name':"code"})
    nameList4 =  bs.find_all('span',attrs={'class':"price-old"})
    nameList2 =  bs.find_all('span',{'class':"price"})
    nameList3 =  bs.find_all('span',attrs={'class':"description"})

    for i in range(len(nameList1)):
        a.append(nameList1[i]['value'])
        b.append(nameList2[i].get_text())
        c.append(nameList3[i].get_text())

       

    for i in range(len(nameList1)):
        print(a[i],' ', b[i],'    ',c[i])

       

    for ii in range(len(nameList1)+ii):
        sheet.cell(row= 2+ii, column=1).value= a[ii]
        sheet.cell(row= 2+ii, column=2).value= c[ii]
        sheet.cell(row= 2+ii, column=3).value= b[ii]
        sheet.cell(row= 2+ii, column=4).value= ad
        ii=ii+1

wb.save("C:\skrypty\excel2.xlsx")

       